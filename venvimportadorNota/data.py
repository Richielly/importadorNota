import pandas as pd
import openpyxl

# 'C:\\Users\\richielly.carvalho\Downloads\\Notas fiscais.xlsx'

class DataSheet():

    def __init__(self, diretory):
        self.diretory = diretory

    def load_data_csv(self, data, sheet, nrows):
        result = pd.read_csv(data, encoding = 'latin1', nrows=nrows)
        return result

    def load_data_excel(self, data, sheet, nrows):
        result = pd.read_excel(data, sheet_name=sheet, nrows=nrows)
        return result

    def load_workbook(self):
        wb = openpyxl.load_workbook(filename=self.diretory)
        return wb

    def load_name_sheets(self):

        wb=DataSheet(self.diretory)
        wb = wb.load_workbook()
        print(wb)
        return wb.sheetnames

    def load_in_sheet(self, sheet_name):

        wb=load_workbook()
        all_sheet = []
        for sheet in wb[sheet_name].iter_rows(values_only=True):
            all_sheet.append(sheet)

        return all_sheet

    def merge_sheets(self, sheet_1, sheet_2, junction, column_consult):   #quando temos colunas com nomes iguais nas duas tabelas como referencia
        merge = pd.merge(tabela_1, tabela_2, how=junction, on=column_consult)
        return merge

